"""
Excel Export Module
Creates beautifully formatted Excel files with enriched contact data
"""

import os
from datetime import datetime
from typing import List
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
import colorama
from colorama import Fore, Style

from config.config import EXPORTS_DIR, EXCEL_FILENAME_TEMPLATE, DEFAULT_SHEET_NAME

class ExcelExporter:
    """Creates professional Excel exports with enriched contact data"""
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
        
        # Define color scheme
        self.colors = {
            'header': 'FF2E75B6',      # Blue
            'high_confidence': 'FF90EE90',  # Light Green
            'medium_confidence': 'FFFFD700', # Gold
            'low_confidence': 'FFFFCCCB',   # Light Red
            'border': 'FF000000'            # Black
        }
    
    def export_contacts(self, contacts: List, filename: str = None) -> str:
        """Export contacts to a formatted Excel file"""
        
        if not contacts:
            print(f"{Fore.YELLOW}⚠️ No contacts to export{Style.RESET_ALL}")
            return None
        
        print(f"{Fore.CYAN}📊 Exporting {len(contacts)} contacts to Excel...{Style.RESET_ALL}")
        
        # Generate filename if not provided
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = EXCEL_FILENAME_TEMPLATE.format(timestamp=timestamp)
        
        filepath = EXPORTS_DIR / filename
        
        # Convert contacts to DataFrame
        df = self._contacts_to_dataframe(contacts)
        
        # Create Excel workbook with formatting
        self._create_formatted_excel(df, filepath)
        
        print(f"{Fore.GREEN}✅ Excel file exported: {filepath}{Style.RESET_ALL}")
        print(f"{Fore.CYAN}📁 File size: {self._get_file_size(filepath)}{Style.RESET_ALL}")
        
        return str(filepath)
    
    def _contacts_to_dataframe(self, contacts: List) -> pd.DataFrame:
        """Convert contact objects to pandas DataFrame"""
        
        data = []
        for contact in contacts:
            row = {
                'Name': contact.name or 'Unknown',
                'Email': contact.email,
                'Location': contact.location or 'Unknown',
                'Estimated Net Worth': contact.estimated_net_worth or 'Unknown',
                'Data Source': contact.data_source or 'None',
                'Confidence Score': f"{contact.confidence:.1f}" if hasattr(contact, 'confidence') else '0.0',
                'Email Frequency': getattr(contact, 'frequency', 1),
                'First Seen': getattr(contact, 'first_seen', datetime.now()).strftime('%Y-%m-%d'),
                'Last Seen': getattr(contact, 'last_seen', datetime.now()).strftime('%Y-%m-%d')
            }
            data.append(row)
        
        return pd.DataFrame(data)
    
    def _create_formatted_excel(self, df: pd.DataFrame, filepath: str):
        """Create Excel file with professional formatting"""
        
        # Create workbook and worksheet
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = DEFAULT_SHEET_NAME
        
        # Write data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            self.worksheet.append(r)
        
        # Apply formatting
        self._format_headers()
        self._format_data_cells()
        self._adjust_column_widths()
        self._add_conditional_formatting()
        self._add_summary_section(df)
        
        # Save workbook
        self.workbook.save(filepath)
    
    def _format_headers(self):
        """Format header row"""
        header_font = Font(
            name='Calibri',
            size=12,
            bold=True,
            color='FFFFFF'
        )
        
        header_fill = PatternFill(
            start_color=self.colors['header'],
            end_color=self.colors['header'],
            fill_type='solid'
        )
        
        header_border = Border(
            left=Side(style='thin', color=self.colors['border']),
            right=Side(style='thin', color=self.colors['border']),
            top=Side(style='thin', color=self.colors['border']),
            bottom=Side(style='thin', color=self.colors['border'])
        )
        
        header_alignment = Alignment(
            horizontal='center',
            vertical='center'
        )
        
        # Apply to header row
        for cell in self.worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = header_border
            cell.alignment = header_alignment
    
    def _format_data_cells(self):
        """Format data cells"""
        data_font = Font(
            name='Calibri',
            size=11
        )
        
        data_border = Border(
            left=Side(style='thin', color='FFCCCCCC'),
            right=Side(style='thin', color='FFCCCCCC'),
            top=Side(style='thin', color='FFCCCCCC'),
            bottom=Side(style='thin', color='FFCCCCCC')
        )
        
        # Apply to all data cells
        for row in self.worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font
                cell.border = data_border
                
                # Center align numeric columns
                if cell.column in [6, 7]:  # Confidence Score, Email Frequency
                    cell.alignment = Alignment(horizontal='center')
    
    def _adjust_column_widths(self):
        """Auto-adjust column widths"""
        column_widths = {
            'A': 25,  # Name
            'B': 35,  # Email
            'C': 25,  # Location
            'D': 20,  # Net Worth
            'E': 15,  # Data Source
            'F': 12,  # Confidence
            'G': 12,  # Frequency
            'H': 12,  # First Seen
            'I': 12   # Last Seen
        }
        
        for column, width in column_widths.items():
            self.worksheet.column_dimensions[column].width = width
    
    def _add_conditional_formatting(self):
        """Add conditional formatting based on confidence scores"""
        
        # Get the range for confidence scores (column F, excluding header)
        max_row = self.worksheet.max_row
        confidence_range = f'F2:F{max_row}'
        
        # Color scale rule for confidence scores
        color_scale_rule = ColorScaleRule(
            start_type='num', start_value=0, start_color='FFFF0000',  # Red for low
            mid_type='num', mid_value=0.5, mid_color='FFFFFF00',     # Yellow for medium
            end_type='num', end_value=1, end_color='FF00FF00'        # Green for high
        )
        
        self.worksheet.conditional_formatting.add(confidence_range, color_scale_rule)
        
        # Data bars for email frequency
        frequency_range = f'G2:G{max_row}'
        data_bar_rule = DataBarRule(
            start_type='min', end_type='max',
            color='FF5B9BD5'
        )
        
        self.worksheet.conditional_formatting.add(frequency_range, data_bar_rule)
    
    def _add_summary_section(self, df: pd.DataFrame):
        """Add summary statistics section"""
        
        # Start summary section after data with some spacing
        summary_start_row = self.worksheet.max_row + 3
        
        # Summary header
        self.worksheet.cell(
            row=summary_start_row, 
            column=1, 
            value="📊 SUMMARY STATISTICS"
        ).font = Font(size=14, bold=True, color=self.colors['header'])
        
        # Calculate statistics
        total_contacts = len(df)
        with_location = len(df[df['Location'] != 'Unknown'])
        with_net_worth = len(df[df['Estimated Net Worth'] != 'Unknown'])
        avg_confidence = df['Confidence Score'].str.replace('', '0').astype(float).mean()
        
        # Add statistics
        stats = [
            ("Total Contacts:", total_contacts),
            ("With Location Data:", f"{with_location} ({with_location/total_contacts*100:.1f}%)"),
            ("With Net Worth Data:", f"{with_net_worth} ({with_net_worth/total_contacts*100:.1f}%)"),
            ("Average Confidence:", f"{avg_confidence:.2f}"),
            ("Export Date:", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        ]
        
        for i, (label, value) in enumerate(stats):
            row = summary_start_row + 2 + i
            self.worksheet.cell(row=row, column=1, value=label).font = Font(bold=True)
            self.worksheet.cell(row=row, column=2, value=str(value))
    
    def _get_file_size(self, filepath: str) -> str:
        """Get human-readable file size"""
        size_bytes = os.path.getsize(filepath)
        
        if size_bytes < 1024:
            return f"{size_bytes} bytes"
        elif size_bytes < 1024 * 1024:
            return f"{size_bytes / 1024:.1f} KB"
        else:
            return f"{size_bytes / (1024 * 1024):.1f} MB"
    
    def create_template_file(self) -> str:
        """Create an empty template file for reference"""
        template_filename = "contact_enrichment_template.xlsx"
        template_path = EXPORTS_DIR / template_filename
        
        # Create sample data
        sample_data = {
            'Name': ['John Doe', 'Jane Smith', 'Bob Johnson'],
            'Email': ['john@example.com', 'jane@company.com', 'bob@startup.io'],
            'Location': ['San Francisco, CA', 'New York, NY', 'Austin, TX'],
            'Estimated Net Worth': ['$250K - $500K', '$500K - $1M', '$1M - $2.5M'],
            'Data Source': ['Demo Data', 'Demo Data', 'Demo Data'],
            'Confidence Score': ['0.8', '0.9', '0.7'],
            'Email Frequency': [5, 12, 3],
            'First Seen': ['2024-01-15', '2024-02-01', '2024-03-10'],
            'Last Seen': ['2024-06-25', '2024-06-28', '2024-06-30']
        }
        
        df = pd.DataFrame(sample_data)
        self._create_formatted_excel(df, template_path)
        
        return str(template_path)

# Demo function
def demo_excel_export():
    """Demo the Excel export functionality"""
    print(f"{Fore.MAGENTA}🚀 Excel Export Demo{Style.RESET_ALL}")
    print("=" * 40)
    
    # Create sample contacts
    from gmail_extractor import Contact
    from enrichment import ContactEnricher
    
    sample_contacts = [
        Contact("John Smith", "john.smith@google.com"),
        Contact("Sarah Johnson", "sarah.j@microsoft.com"),
        Contact("Mike Chen", "mike.chen@university.edu"),
        Contact("Lisa Rodriguez", "l.rodriguez@salesforce.com"),
        Contact("David Kim", "david@startup.com")
    ]
    
    # Add some sample enrichment data
    for i, contact in enumerate(sample_contacts):
        contact.location = ["San Francisco, CA", "Seattle, WA", "Boston, MA", "Austin, TX", "New York, NY"][i]
        contact.estimated_net_worth = ["$250K - $500K", "$500K - $1M", "$100K - $250K", "$1M - $2.5M", "$2.5M - $5M"][i]
        contact.data_source = "Demo Data"
        contact.confidence = [0.9, 0.8, 0.6, 0.95, 0.7][i]
        contact.frequency = [10, 25, 5, 15, 8][i]
    
    # Export to Excel
    exporter = ExcelExporter()
    filepath = exporter.export_contacts(sample_contacts, "demo_contacts.xlsx")
    
    print(f"\n{Fore.GREEN}✅ Demo export completed!{Style.RESET_ALL}")
    print(f"{Fore.CYAN}📁 File location: {filepath}{Style.RESET_ALL}")

if __name__ == "__main__":
    demo_excel_export()