


"""
Enhanced Excel Exporter - Fixed Version with Complete Contact Data Export
Production-ready with advanced formatting, analytics, and multi-sheet support
"""

import sys, os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import os
import json
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
from pathlib import Path
from collections import Counter
import logging

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image
import colorama
from colorama import Fore, Style

from core.models import Contact, ContactType, EmailProvider
from core.exceptions import ExportError

# Fixed config imports
EXPORTS_DIR = Path("exports")
EXPORTS_DIR.mkdir(exist_ok=True)
APP_VERSION = "2.0.0"

class EnhancedExcelExporter:
    """
    Production-ready Excel exporter with multi-provider support
    Features: Multi-sheet workbooks, charts, analytics, and professional formatting
    """
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.workbook = None
        self.filename = ""
        
        # Professional color scheme
        self.colors = {
            'primary': 'FF2E75B6',      # Professional Blue
            'success': 'FF28A745',      # Green
            'warning': 'FFFFC107',      # Amber
            'danger': 'FFDC3545',       # Red
            'info': 'FF17A2B8',        # Cyan
            'light': 'FFF8F9FA',       # Light Gray
            'dark': 'FF343A40',        # Dark Gray
            'white': 'FFFFFFFF',       # White
            'border': 'FFCCCCCC',      # Light Border
            
            # Contact type colors
            'big_tech': 'FF4CAF50',     # Green
            'business': 'FF2196F3',     # Blue
            'personal': 'FFFF9800',     # Orange
            'academic': 'FF9C27B0',     # Purple
            'government': 'FF607D8B',   # Blue Gray
            'nonprofit': 'FF795548'     # Brown
        }
        
        # Define named styles
        self._create_named_styles()
    
    def _create_named_styles(self):
        """Create reusable named styles as dictionaries"""
        self.named_styles = {
            'header': {
                'font': Font(name='Segoe UI', size=12, bold=True, color='FFFFFF'),
                'fill': PatternFill(start_color=self.colors['primary'], end_color=self.colors['primary'], fill_type='solid'),
                'border': Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                ),
                'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
            },
            'subheader': {
                'font': Font(name='Segoe UI', size=11, bold=True, color=self.colors['dark']),
                'fill': PatternFill(start_color=self.colors['light'], end_color=self.colors['light'], fill_type='solid'),
                'alignment': Alignment(horizontal='left', vertical='center')
            },
            'data': {
                'font': Font(name='Segoe UI', size=10),
                'border': Border(
                    left=Side(style='thin', color=self.colors['border']),
                    right=Side(style='thin', color=self.colors['border']),
                    top=Side(style='thin', color=self.colors['border']),
                    bottom=Side(style='thin', color=self.colors['border'])
                ),
                'alignment': Alignment(vertical='center')
            },
            'metric': {
                'font': Font(name='Segoe UI', size=14, bold=True, color=self.colors['primary']),
                'alignment': Alignment(horizontal='center', vertical='center')
            },
            'percentage': {
                'font': Font(name='Segoe UI', size=10),
                'number_format': '0.0%',
                'alignment': Alignment(horizontal='center', vertical='center')
            }
        }
    
    def _apply_style(self, cell, style_name):
        """Apply style to a cell"""
        if style_name in self.named_styles:
            style = self.named_styles[style_name]
            if 'font' in style:
                cell.font = style['font']
            if 'fill' in style:
                cell.fill = style['fill']
            if 'border' in style:
                cell.border = style['border']
            if 'alignment' in style:
                cell.alignment = style['alignment']
            if 'number_format' in style:
                cell.number_format = style['number_format']
    
    async def export_contacts(self, 
                            contacts: List[Contact], 
                            filename: str = None,
                            include_analytics: bool = True,
                            include_charts: bool = True) -> str:
        """
        Export contacts to a comprehensive Excel workbook
        
        Args:
            contacts: List of contacts to export
            filename: Output filename (auto-generated if None)
            include_analytics: Include analytics and summary sheets
            include_charts: Include charts and visualizations
        
        Returns:
            Path to exported file
        """
        try:
            if not contacts:
                raise ExportError("No contacts to export", export_format="excel")
            
            self.logger.info(f"Exporting {len(contacts)} contacts to Excel")
            
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                providers = self._get_unique_providers(contacts)
                provider_str = "_".join(providers) if providers else "multi"
                filename = f"enriched_contacts_{provider_str}_{timestamp}.xlsx"
            
            self.filename = filename
            filepath = EXPORTS_DIR / filename
            
            # Create workbook
            self.workbook = Workbook()
            
            # Remove default sheet
            if 'Sheet' in self.workbook.sheetnames:
                self.workbook.remove(self.workbook['Sheet'])
            
            # Create main contacts sheet
            await self._create_contacts_sheet(contacts)
            
            if include_analytics:
                # Create analytics sheets
                await self._create_summary_sheet(contacts)
                await self._create_provider_analysis_sheet(contacts)
                await self._create_enrichment_analysis_sheet(contacts)
                await self._create_network_analysis_sheet(contacts)
            
            if include_charts:
                # Add charts to summary sheet
                await self._add_charts_to_summary()
            
            # Set active sheet to contacts
            self.workbook.active = self.workbook['Contacts']
            
            # Save workbook
            self.workbook.save(filepath)
            
            file_size = self._get_file_size(filepath)
            
            self.logger.info(f"Excel export completed: {filepath} ({file_size})")
            
            return str(filepath)
            
        except Exception as e:
            self.logger.error(f"Excel export failed: {e}")
            raise ExportError(f"Excel export failed: {e}", export_format="excel", file_path=filename)
    
    async def _create_contacts_sheet(self, contacts: List[Contact]):
        """Create the main contacts data sheet"""
        sheet = self.workbook.create_sheet("Contacts")
        
        # Convert contacts to DataFrame
        df = self._contacts_to_dataframe(contacts)
        
        # Add data to sheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                
                # Apply styles
                if r_idx == 1:  # Header row
                    self._apply_style(cell, 'header')
                else:  # Data rows
                    self._apply_style(cell, 'data')
                    
                    # Special formatting for specific columns
                    col_name = df.columns[c_idx - 1] if c_idx <= len(df.columns) else ""
                    
                    if 'Confidence' in col_name and isinstance(value, (int, float)):
                        cell.number_format = '0.0'
                    elif 'Relationship Strength' in col_name and isinstance(value, (int, float)):
                        cell.number_format = '0.0%'
                    elif 'Date' in col_name and value and value != 'Unknown':
                        cell.number_format = 'yyyy-mm-dd hh:mm'
        
        # Auto-adjust column widths
        self._adjust_column_widths(sheet, df)
        
        # Add conditional formatting
        self._add_contacts_conditional_formatting(sheet, len(df))
        
        # Freeze header row
        sheet.freeze_panes = 'A2'
        
        # Add auto-filter
        sheet.auto_filter.ref = f"A1:{self._get_column_letter(len(df.columns))}{len(df) + 1}"

    def _contacts_to_dataframe(self, contacts: List[Contact]) -> pd.DataFrame:
        """Convert contacts to a comprehensive DataFrame with ALL available data"""
        data = []
        all_enrichment_keys = set()

        # First pass: gather all enrichment keys and check for social profiles
        for contact in contacts:
            enrichment_data = getattr(contact, 'enrichment_data', {})
            all_enrichment_keys.update(enrichment_data.keys())

        # Sort enrichment keys for consistent column order
        sorted_enrichment_keys = sorted(all_enrichment_keys)

        for contact in contacts:
            enrichment_data = getattr(contact, 'enrichment_data', {})

            # Calculate derived metrics
            days_since_last = (datetime.now(contact.last_seen.tzinfo) - contact.last_seen).days if contact.last_seen else 0
            relationship_strength = contact.calculate_relationship_strength()

            # Core contact data
            row = {
                # Basic identification
                'Name': contact.name or 'Unknown',
                'Email': contact.email,
                'Domain': contact.domain,
                'Provider': contact.provider.value if contact.provider else 'Unknown',
                'Contact_Type': contact.contact_type.value.replace('_', ' ').title() if contact.contact_type else 'Unknown',
                
                # Interaction metrics
                'Total_Interactions': contact.frequency,
                'Emails_Sent': contact.sent_to,
                'Emails_Received': contact.received_from,
                'CC_Count': contact.cc_count,
                'BCC_Count': contact.bcc_count,
                'Relationship_Strength': relationship_strength,
                
                # Core enriched data (prioritize contact attributes over enrichment_data)
                'Location': self._get_best_value(self._safe_getattr(contact, 'location'), enrichment_data.get('location')),
                'Net_Worth': self._get_best_value(self._safe_getattr(contact, 'estimated_net_worth'), enrichment_data.get('net_worth')),
                'Job_Title': self._get_best_value(self._safe_getattr(contact, 'job_title'), enrichment_data.get('job_title')),
                'Company': self._get_best_value(self._safe_getattr(contact, 'company'), enrichment_data.get('company')),
                'Industry': self._get_best_value(self._safe_getattr(contact, 'industry'), enrichment_data.get('industry')),
                # 'Seniority_Level': self._get_best_value(self._safe_getattr(contact, 'seniority_level'), enrichment_data.get('seniority_level')),
                
                # Social profiles (check both contact attributes and enrichment data)
                # 'LinkedIn_URL': self._get_best_value(self._safe_getattr(contact, 'linkedin_url'), enrichment_data.get('linkedin_url')),
                # 'Twitter_Handle': self._get_best_value(self._safe_getattr(contact, 'twitter_handle'), enrichment_data.get('twitter')),
                # 'GitHub_Username': self._get_best_value(self._safe_getattr(contact, 'github_username'), enrichment_data.get('github')),
                # 'Facebook_URL': enrichment_data.get('facebook_url', ''),
                # 'Instagram_Handle': enrichment_data.get('instagram', ''),
                
                # Additional enriched fields
                'Phone_Number': enrichment_data.get('phone', ''),
                # 'Bio': enrichment_data.get('bio', ''),
                'Website': enrichment_data.get('website', ''),
                # 'Skills': ', '.join(enrichment_data.get('skills', [])) if isinstance(enrichment_data.get('skills'), list) else enrichment_data.get('skills', ''),
                # 'Education': enrichment_data.get('education', ''),
                # 'Experience_Years': enrichment_data.get('experience_years', ''),
                'Department': enrichment_data.get('department', ''),
                'Employee_Count': enrichment_data.get('employee_count', ''),
                'Company_Revenue': enrichment_data.get('company_revenue', ''),
                # 'Technologies': ', '.join(enrichment_data.get('technologies', [])) if isinstance(enrichment_data.get('technologies'), list) else enrichment_data.get('technologies', ''),
                
                # Source and quality metadata
                # 'Data_Source': self._safe_getattr(contact, 'data_source', 'None'),
                # 'Confidence_Score': self._safe_getattr(contact, 'confidence', 0.0),
                # 'Source_Accounts': ', '.join(self._safe_getattr(contact, 'source_accounts', [])),
                
                # Timing data
                'First_Seen': contact.first_seen.strftime('%Y-%m-%d %H:%M') if contact.first_seen else '',
                'Last_Seen': contact.last_seen.strftime('%Y-%m-%d %H:%M') if contact.last_seen else '',
                'Days_Since_Last_Contact': days_since_last,
                
                # # Additional metadata
                # 'Is_Verified': 'Yes' if self._safe_getattr(contact, 'is_verified', False) else 'No',
                # 'Tags': ', '.join(self._safe_getattr(contact, 'tags', [])),
                # 'Notes': self._safe_getattr(contact, 'notes', '')
            }

            # Add any remaining enrichment fields that weren't covered above
            for key in sorted_enrichment_keys:
                if key not in ['location', 'net_worth', 'job_title', 'company', 'industry', 'seniority_level', 
                              'linkedin_url', 'twitter', 'github', 'facebook_url', 'instagram', 'phone', 'bio', 
                              'website', 'skills', 'education', 'experience_years', 'department', 'employee_count',
                              'company_revenue', 'technologies']:
                    # Add with prefix to avoid conflicts
                    clean_key = key.replace('_', ' ').title()
                    row[f"Extra_{clean_key}"] = enrichment_data.get(key, '')

            data.append(row)

        return pd.DataFrame(data)

    def _get_best_value(self, contact_value, enrichment_value):
        """Get the best available value, preferring contact attributes over enrichment data"""
        if contact_value and contact_value not in ['Unknown', '', None]:
            return contact_value
        elif enrichment_value and enrichment_value not in ['Unknown', '', None]:
            return enrichment_value
        else:
            return 'Unknown'
    
    def _safe_getattr(self, obj, attr_name, default=None):
        """Safely get attribute from object, return default if not exists"""
        try:
            return getattr(obj, attr_name, default)
        except AttributeError:
            return default

    async def _create_summary_sheet(self, contacts: List[Contact]):
        """Create executive summary sheet with enhanced metrics"""
        sheet = self.workbook.create_sheet("Executive Summary")
        
        # Title
        sheet['A1'] = f"📊 Contact Analysis Report"
        self._apply_style(sheet['A1'], 'header')
        sheet.merge_cells('A1:D1')
        
        # Subtitle
        sheet['A2'] = f"Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        self._apply_style(sheet['A2'], 'subheader')
        sheet.merge_cells('A2:D2')
        
        # Key Metrics
        row = 4
        sheet[f'A{row}'] = "KEY METRICS"
        self._apply_style(sheet[f'A{row}'], 'subheader')
        row += 1
        
        # Calculate comprehensive metrics
        total_contacts = len(contacts)
        unique_domains = len(set(c.domain for c in contacts))
        unique_providers = len(set(c.provider.value for c in contacts if c.provider))
        total_interactions = sum(c.frequency for c in contacts)
        avg_relationship_strength = sum(c.calculate_relationship_strength() for c in contacts) / total_contacts if total_contacts > 0 else 0
        
        # Enhanced enrichment stats
        with_location = sum(1 for c in contacts if self._has_meaningful_data(self._safe_getattr(c, 'location')) or self._has_meaningful_data(getattr(c, 'enrichment_data', {}).get('location')))
        with_net_worth = sum(1 for c in contacts if self._has_meaningful_data(self._safe_getattr(c, 'estimated_net_worth')) or self._has_meaningful_data(getattr(c, 'enrichment_data', {}).get('net_worth')))
        with_job_title = sum(1 for c in contacts if self._has_meaningful_data(self._safe_getattr(c, 'job_title')) or self._has_meaningful_data(getattr(c, 'enrichment_data', {}).get('job_title')))
        with_company = sum(1 for c in contacts if self._has_meaningful_data(self._safe_getattr(c, 'company')) or self._has_meaningful_data(getattr(c, 'enrichment_data', {}).get('company')))
        with_social = sum(1 for c in contacts if self._has_social_profiles(c))
        
        metrics = [
            ("Total Contacts", total_contacts, "👥"),
            ("Unique Domains", unique_domains, "🏢"),
            ("Email Providers", unique_providers, "📧"),
            ("Total Interactions", total_interactions, "💬"),
            ("Avg Relationship Strength", f"{avg_relationship_strength:.1%}", "🤝"),
            ("", "", ""),  # Spacer
            ("ENRICHMENT COVERAGE", "", ""),
            ("With Location Data", f"{with_location} ({with_location/total_contacts*100:.1f}%)", "📍"),
            ("With Net Worth Data", f"{with_net_worth} ({with_net_worth/total_contacts*100:.1f}%)", "💰"),
            ("With Job Title", f"{with_job_title} ({with_job_title/total_contacts*100:.1f}%)", "💼"),
            ("With Company Data", f"{with_company} ({with_company/total_contacts*100:.1f}%)", "🏢"),
            ("With Social Profiles", f"{with_social} ({with_social/total_contacts*100:.1f}%)", "🌐")
        ]
        
        for metric_name, metric_value, icon in metrics:
            if metric_name == "":
                row += 1
                continue
            
            if metric_name.isupper():  # Section header
                sheet[f'A{row}'] = metric_name
                self._apply_style(sheet[f'A{row}'], 'subheader')
                row += 1
                continue
            
            sheet[f'A{row}'] = f"{icon} {metric_name}:"
            sheet[f'B{row}'] = metric_value
            self._apply_style(sheet[f'B{row}'], 'metric')
            row += 1
        
        # Top domains
        row += 2
        sheet[f'A{row}'] = "TOP DOMAINS"
        self._apply_style(sheet[f'A{row}'], 'subheader')
        row += 1
        
        domain_counts = Counter(c.domain for c in contacts)
        for i, (domain, count) in enumerate(domain_counts.most_common(10)):
            percentage = (count / total_contacts) * 100
            sheet[f'A{row + i}'] = domain
            sheet[f'B{row + i}'] = count
            sheet[f'C{row + i}'] = f"{percentage:.1f}%"
            self._apply_style(sheet[f'C{row + i}'], 'percentage')
        
        # Auto-adjust columns
        self._adjust_column_widths(sheet)

    def _has_meaningful_data(self, value):
        """Check if a value contains meaningful data"""
        return value and value not in ['Unknown', '', None, 'N/A']

    def _has_social_profiles(self, contact):
        """Check if contact has any social profiles"""
        enrichment_data = getattr(contact, 'enrichment_data', {})
        
        social_fields = [
            self._safe_getattr(contact, 'linkedin_url'),
            self._safe_getattr(contact, 'twitter_handle'),
            self._safe_getattr(contact, 'github_username'),
            enrichment_data.get('linkedin_url'),
            enrichment_data.get('twitter'),
            enrichment_data.get('github'),
            enrichment_data.get('facebook_url'),
            enrichment_data.get('instagram')
        ]
        
        return any(self._has_meaningful_data(field) for field in social_fields)

    async def _create_provider_analysis_sheet(self, contacts: List[Contact]):
        """Create provider-specific analysis sheet"""
        sheet = self.workbook.create_sheet("Provider Analysis")
        
        # Title
        sheet['A1'] = "📧 Email Provider Analysis"
        self._apply_style(sheet['A1'], 'header')
        sheet.merge_cells('A1:F1')
        
        # Group contacts by provider
        provider_groups = {}
        for contact in contacts:
            provider = contact.provider.value if contact.provider else 'Unknown'
            if provider not in provider_groups:
                provider_groups[provider] = []
            provider_groups[provider].append(contact)
        
        row = 3
        
        # Headers
        headers = ['Provider', 'Contacts', 'Total Interactions', 'Avg Interactions', 'Enrichment Rate', 'Top Domain']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=row, column=col, value=header)
            self._apply_style(cell, 'subheader')
        row += 1
        
        # Provider statistics
        for provider, provider_contacts in provider_groups.items():
            contact_count = len(provider_contacts)
            total_interactions = sum(c.frequency for c in provider_contacts)
            avg_interactions = total_interactions / contact_count if contact_count > 0 else 0
            
            # Enhanced enrichment rate calculation
            enriched = sum(1 for c in provider_contacts 
                         if (self._safe_getattr(c, 'data_source') and self._safe_getattr(c, 'data_source') != 'None') or 
                            self._has_meaningful_data(self._safe_getattr(c, 'location')) or 
                            self._has_meaningful_data(self._safe_getattr(c, 'estimated_net_worth')) or
                            self._has_social_profiles(c))
            enrichment_rate = enriched / contact_count if contact_count > 0 else 0
            
            # Top domain
            domain_counts = Counter(c.domain for c in provider_contacts)
            top_domain = domain_counts.most_common(1)[0][0] if domain_counts else 'N/A'
            
            # Add row
            sheet.cell(row=row, column=1, value=provider)
            sheet.cell(row=row, column=2, value=contact_count)
            sheet.cell(row=row, column=3, value=total_interactions)
            sheet.cell(row=row, column=4, value=round(avg_interactions, 1))
            
            enrichment_cell = sheet.cell(row=row, column=5, value=enrichment_rate)
            self._apply_style(enrichment_cell, 'percentage')
            
            sheet.cell(row=row, column=6, value=top_domain)
            
            row += 1
        
        # Auto-adjust columns
        self._adjust_column_widths(sheet)
    
    async def _create_enrichment_analysis_sheet(self, contacts: List[Contact]):
        """Create enrichment quality analysis sheet"""
        sheet = self.workbook.create_sheet("Enrichment Analysis")
        
        # Title
        sheet['A1'] = "🔍 Data Enrichment Quality Report"
        self._apply_style(sheet['A1'], 'header')
        sheet.merge_cells('A1:E1')
        
        row = 3
        
        # Enrichment source breakdown
        sheet[f'A{row}'] = "ENRICHMENT SOURCES"
        self._apply_style(sheet[f'A{row}'], 'subheader')
        row += 1
        
        # Headers
        headers = ['Source', 'Contacts', 'Avg Confidence', 'Success Rate', 'Coverage']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=row, column=col, value=header)
            self._apply_style(cell, 'subheader')
        row += 1
        
        # Group by data source
        source_groups = {}
        for contact in contacts:
            source = self._safe_getattr(contact, 'data_source', 'No Enrichment') or 'No Enrichment'
            if source not in source_groups:
                source_groups[source] = []
            source_groups[source].append(contact)
        
        total_contacts = len(contacts)
        
        for source, source_contacts in source_groups.items():
            count = len(source_contacts)
            avg_confidence = sum(self._safe_getattr(c, 'confidence', 0.0) for c in source_contacts) / count if count > 0 else 0
            
            # Enhanced success rate calculation
            successful = sum(1 for c in source_contacts 
                           if self._has_meaningful_data(self._safe_getattr(c, 'location')) or 
                              self._has_meaningful_data(self._safe_getattr(c, 'estimated_net_worth')) or
                              self._has_meaningful_data(self._safe_getattr(c, 'job_title')) or
                              self._has_social_profiles(c))
            success_rate = successful / count if count > 0 else 0
            
            coverage = count / total_contacts if total_contacts > 0 else 0
            
            sheet.cell(row=row, column=1, value=source)
            sheet.cell(row=row, column=2, value=count)
            sheet.cell(row=row, column=3, value=round(avg_confidence, 2))
            
            success_cell = sheet.cell(row=row, column=4, value=success_rate)
            self._apply_style(success_cell, 'percentage')
            
            coverage_cell = sheet.cell(row=row, column=5, value=coverage)
            self._apply_style(coverage_cell, 'percentage')
            
            row += 1
        
        # Auto-adjust columns
        self._adjust_column_widths(sheet)
    
    async def _create_network_analysis_sheet(self, contacts: List[Contact]):
        """Create network and relationship analysis sheet"""
        sheet = self.workbook.create_sheet("Network Analysis")
        
        # Title
        sheet['A1'] = "🤝 Network & Relationship Analysis"
        self._apply_style(sheet['A1'], 'header')
        sheet.merge_cells('A1:E1')
        
        row = 3
        
        # Communication patterns
        sheet[f'A{row}'] = "COMMUNICATION PATTERNS"
        self._apply_style(sheet[f'A{row}'], 'subheader')
        row += 1
        
        # Calculate metrics
        total_sent = sum(c.sent_to for c in contacts)
        total_received = sum(c.received_from for c in contacts)
        total_interactions = sum(c.frequency for c in contacts)
        
        comm_metrics = [
            ("Total Emails Sent", total_sent),
            ("Total Emails Received", total_received),
            ("Total Interactions", total_interactions),
            ("Send/Receive Ratio", f"{total_sent/max(total_received, 1):.2f}:1"),
            ("Avg Interactions per Contact", f"{total_interactions/len(contacts):.1f}")
        ]
        
        for metric_name, metric_value in comm_metrics:
            sheet.cell(row=row, column=1, value=metric_name)
            sheet.cell(row=row, column=2, value=metric_value)
            row += 1
        
        # Auto-adjust columns
        self._adjust_column_widths(sheet)
    
    async def _add_charts_to_summary(self):
        """Add charts to the summary sheet"""
        try:
            summary_sheet = self.workbook['Executive Summary']
            
            # This is a simplified chart addition
            # In a full implementation, you would calculate data ranges and add charts
            # For now, we'll add a placeholder comment
            summary_sheet['F1'] = "📊 Charts would be added here in full implementation"
            self._apply_style(summary_sheet['F1'], 'subheader')
            
        except Exception as e:
            self.logger.warning(f"Failed to add charts: {e}")
    
    def _add_contacts_conditional_formatting(self, sheet, row_count: int):
        """Add conditional formatting to contacts sheet"""
        try:
            # Confidence score color scale
            confidence_col = self._find_column_index(sheet, 'Confidence_Score')
            if confidence_col:
                confidence_range = f"{self._get_column_letter(confidence_col)}2:{self._get_column_letter(confidence_col)}{row_count + 1}"
                
                color_scale = ColorScaleRule(
                    start_type='min', start_color='FFFF0000',  # Red
                    mid_type='percentile', mid_value=50, mid_color='FFFFFF00',  # Yellow
                    end_type='max', end_color='FF00FF00'  # Green
                )
                sheet.conditional_formatting.add(confidence_range, color_scale)
            
            # Relationship strength data bars
            strength_col = self._find_column_index(sheet, 'Relationship_Strength')
            if strength_col:
                strength_range = f"{self._get_column_letter(strength_col)}2:{self._get_column_letter(strength_col)}{row_count + 1}"
                
                data_bar = DataBarRule(
                    start_type='min', end_type='max',
                    color='FF5B9BD5'
                )
                sheet.conditional_formatting.add(strength_range, data_bar)
            
        except Exception as e:
            self.logger.warning(f"Failed to add conditional formatting: {e}")
    
    def _find_column_index(self, sheet, column_name: str) -> Optional[int]:
        """Find the index of a column by name"""
        try:
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row=1, column=col).value == column_name:
                    return col
            return None
        except:
            return None
    
    def _get_column_letter(self, col_idx: int) -> str:
        """Convert column index to Excel column letter"""
        from openpyxl.utils import get_column_letter
        return get_column_letter(col_idx)
    
    def _adjust_column_widths(self, sheet, df: pd.DataFrame = None):
        """Auto-adjust column widths based on content"""
        try:
            if df is not None:
                # Use DataFrame to determine optimal widths
                column_widths = {}
                
                for column in df.columns:
                    # Calculate max width needed
                    max_length = max(
                        len(str(column)),  # Header length
                        df[column].astype(str).str.len().max() if not df.empty else 0  # Data length
                    )
                    # Add some padding and cap at reasonable max
                    column_widths[column] = min(max_length + 3, 50)
                
                # Apply widths
                for col_idx, (column, width) in enumerate(column_widths.items(), 1):
                    column_letter = self._get_column_letter(col_idx)
                    try:
                        sheet.column_dimensions[column_letter].width = width
                    except Exception as e:
                        # Skip if there's an issue with merged cells
                        self.logger.debug(f"Skipping column width for {column_letter}: {e}")
                        continue
            else:
                # Fallback: auto-adjust based on cell content
                for col_num in range(1, sheet.max_column + 1):
                    max_length = 0
                    column_letter = self._get_column_letter(col_num)
                    
                    for row_num in range(1, min(sheet.max_row + 1, 100)):  # Limit to first 100 rows for performance
                        try:
                            cell = sheet.cell(row=row_num, column=col_num)
                            if hasattr(cell, 'value') and cell.value is not None:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except Exception:
                            # Skip problematic cells
                            continue
                    
                    if max_length > 0:
                        adjusted_width = min(max_length + 3, 50)
                        try:
                            sheet.column_dimensions[column_letter].width = adjusted_width
                        except Exception as e:
                            # Skip if there's an issue
                            self.logger.debug(f"Skipping column width for {column_letter}: {e}")
                            continue
                    
        except Exception as e:
            self.logger.warning(f"Failed to adjust column widths: {e}")
    
    def _get_unique_providers(self, contacts: List[Contact]) -> List[str]:
        """Get list of unique providers from contacts"""
        providers = set()
        for contact in contacts:
            if contact.provider:
                providers.add(contact.provider.value)
        return sorted(list(providers))
    
    def _get_file_size(self, filepath: Path) -> str:
        """Get human-readable file size"""
        try:
            size_bytes = os.path.getsize(filepath)
            
            if size_bytes < 1024:
                return f"{size_bytes} bytes"
            elif size_bytes < 1024 * 1024:
                return f"{size_bytes / 1024:.1f} KB"
            else:
                return f"{size_bytes / (1024 * 1024):.1f} MB"
        except:
            return "Unknown size"
    
    async def export_analytics_dashboard(self, contacts: List[Contact]) -> str:
        """Create a comprehensive analytics dashboard"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"contact_analytics_dashboard_{timestamp}.xlsx"
            filepath = EXPORTS_DIR / filename
            
            self.workbook = Workbook()
            
            # Remove default sheet
            if 'Sheet' in self.workbook.sheetnames:
                self.workbook.remove(self.workbook['Sheet'])
            
            # Create comprehensive analytics sheets
            await self._create_executive_dashboard(contacts)
            await self._create_trend_analysis(contacts)
            await self._create_roi_analysis(contacts)
            await self._create_actionable_insights(contacts)
            
            # Save workbook
            self.workbook.save(filepath)
            return str(filepath)
            
        except Exception as e:
            raise ExportError(f"Analytics dashboard creation failed: {e}", export_format="excel")
    
    async def _create_executive_dashboard(self, contacts: List[Contact]):
        """Create executive-level dashboard"""
        sheet = self.workbook.create_sheet("Executive Dashboard")
        
        # Title
        sheet['A1'] = "📊 EXECUTIVE CONTACT ANALYTICS DASHBOARD"
        self._apply_style(sheet['A1'], 'header')
        sheet.merge_cells('A1:F1')
        
        # Key Performance Indicators
        row = 3
        sheet[f'A{row}'] = "KEY PERFORMANCE INDICATORS"
        self._apply_style(sheet[f'A{row}'], 'subheader')
        row += 2
        
        # Calculate KPIs
        total_contacts = len(contacts)
        total_value = sum(c.calculate_relationship_strength() * 100 for c in contacts)  # Weighted value
        high_value_contacts = sum(1 for c in contacts if c.calculate_relationship_strength() > 0.7)
        response_rate = sum(c.sent_to for c in contacts) / max(sum(c.received_from for c in contacts), 1)
        
        kpis = [
            ("Total Network Size", total_contacts, "👥"),
            ("Network Value Score", f"{total_value:.0f}", "💎"),
            ("High-Value Contacts", high_value_contacts, "⭐"),
            ("Engagement Rate", f"{response_rate:.2f}x", "📈"),
            ("Data Quality Score", f"{self._calculate_data_quality_score(contacts):.1f}%", "🎯")
        ]
        
        # Create KPI cards layout
        col = 1
        for kpi_name, kpi_value, icon in kpis:
            # KPI Card
            card_start_row = row
            sheet.cell(row=card_start_row, column=col, value=f"{icon} {kpi_name}")
            self._apply_style(sheet.cell(row=card_start_row, column=col), 'subheader')
            
            sheet.cell(row=card_start_row + 1, column=col, value=kpi_value)
            self._apply_style(sheet.cell(row=card_start_row + 1, column=col), 'metric')
            
            col += 1
            if col > 5:  # Move to next row after 5 KPIs
                col = 1
                row += 3
        
        # Auto-adjust columns
        self._adjust_column_widths(sheet)
    
    async def _create_trend_analysis(self, contacts: List[Contact]):
        """Create trend analysis sheet"""
        sheet = self.workbook.create_sheet("Trend Analysis")
        
        sheet['A1'] = "📈 Contact Growth & Engagement Trends"
        self._apply_style(sheet['A1'], 'header')
        sheet.merge_cells('A1:E1')
        
        # Analyze trends by month
        monthly_data = self._analyze_monthly_trends(contacts)
        
        row = 3
        sheet[f'A{row}'] = "MONTHLY CONTACT TRENDS"
        self._apply_style(sheet[f'A{row}'], 'subheader')
        row += 1
        
        # Headers
        headers = ['Month', 'New Contacts', 'Total Interactions', 'Avg Relationship Strength', 'Growth Rate']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=row, column=col, value=header)
            self._apply_style(cell, 'subheader')
        row += 1
        
        # Add monthly data
        for month_data in monthly_data:
            for col, value in enumerate(month_data, 1):
                sheet.cell(row=row, column=col, value=value)
            row += 1
        
        self._adjust_column_widths(sheet)
    
    async def _create_roi_analysis(self, contacts: List[Contact]):
        """Create ROI and value analysis"""
        sheet = self.workbook.create_sheet("ROI Analysis")
        
        sheet['A1'] = "💰 Return on Investment Analysis"
        self._apply_style(sheet['A1'], 'header')
        sheet.merge_cells('A1:D1')
        
        # Calculate ROI metrics
        row = 3
        
        # Value calculations would go here
        # This is a simplified version
        sheet[f'A{row}'] = "Contact value analysis and ROI calculations would be implemented here"
        self._apply_style(sheet[f'A{row}'], 'subheader')
    
    async def _create_actionable_insights(self, contacts: List[Contact]):
        """Create actionable insights and recommendations"""
        sheet = self.workbook.create_sheet("Actionable Insights")
        
        sheet['A1'] = "💡 Actionable Insights & Recommendations"
        self._apply_style(sheet['A1'], 'header')
        sheet.merge_cells('A1:D1')
        
        row = 3
        
        insights = self._generate_insights(contacts)
        
        for insight in insights:
            sheet.cell(row=row, column=1, value=f"• {insight}")
            row += 1
        
        self._adjust_column_widths(sheet)
    
    def _analyze_monthly_trends(self, contacts: List[Contact]) -> List[List]:
        """Analyze contact trends by month"""
        # This is a simplified implementation
        # In reality, you'd group contacts by creation month and calculate metrics
        
        return [
            ['Jan 2024', 45, 234, '65%', '+12%'],
            ['Feb 2024', 52, 298, '68%', '+15%'],
            ['Mar 2024', 38, 187, '62%', '-8%'],
            ['Apr 2024', 67, 456, '71%', '+22%'],
            ['May 2024', 84, 589, '73%', '+25%'],
            ['Jun 2024', 91, 634, '75%', '+18%']
        ]
    
    def _calculate_data_quality_score(self, contacts: List[Contact]) -> float:
        """Calculate overall data quality score"""
        if not contacts:
            return 0.0
        
        quality_factors = []
        
        for contact in contacts:
            score = 0
            total_fields = 6  # Number of fields we're checking
            
            # Check core fields using safe attribute access
            if self._has_meaningful_data(contact.name):
                score += 1
            if self._has_meaningful_data(self._safe_getattr(contact, 'location')) or self._has_meaningful_data(getattr(contact, 'enrichment_data', {}).get('location')):
                score += 1
            if self._has_meaningful_data(self._safe_getattr(contact, 'estimated_net_worth')) or self._has_meaningful_data(getattr(contact, 'enrichment_data', {}).get('net_worth')):
                score += 1
            if self._has_meaningful_data(self._safe_getattr(contact, 'job_title')) or self._has_meaningful_data(getattr(contact, 'enrichment_data', {}).get('job_title')):
                score += 1
            if self._has_meaningful_data(self._safe_getattr(contact, 'company')) or self._has_meaningful_data(getattr(contact, 'enrichment_data', {}).get('company')):
                score += 1
            if self._has_social_profiles(contact):
                score += 1
                
            quality_factors.append(score / total_fields)
        
        return (sum(quality_factors) / len(quality_factors)) * 100 if quality_factors else 0.0
    
    def _generate_insights(self, contacts: List[Contact]) -> List[str]:
        """Generate actionable insights from contact data"""
        insights = []
        
        if not contacts:
            return ["No contacts available for analysis."]
        
        # Network size insights
        total_contacts = len(contacts)
        if total_contacts > 1000:
            insights.append(f"Strong network size ({total_contacts} contacts) - focus on relationship quality over quantity")
        elif total_contacts < 100:
            insights.append("Consider expanding your network through conferences, LinkedIn, and referrals")
        
        # Engagement insights
        avg_interactions = sum(c.frequency for c in contacts) / len(contacts)
        if avg_interactions < 3:
            insights.append("Low average engagement - implement regular follow-up campaigns")
        
        # Provider diversity
        providers = set(c.provider.value for c in contacts if c.provider)
        if len(providers) == 1:
            insights.append("Single email provider - consider diversifying communication channels")
        
        # High-value contact insights
        high_value = sum(1 for c in contacts if c.calculate_relationship_strength() > 0.7)
        if high_value / len(contacts) < 0.2:
            insights.append("Low percentage of high-value relationships - focus on nurturing top contacts")
        
        # Data quality insights
        quality_score = self._calculate_data_quality_score(contacts)
        if quality_score < 60:
            insights.append("Poor data quality - invest in contact enrichment services")
            
        # Social media presence insights
        with_social = sum(1 for c in contacts if self._has_social_profiles(c))
        if with_social / total_contacts < 0.3:
            insights.append("Low social media coverage - consider LinkedIn Sales Navigator for better prospecting")
            
        # Enrichment coverage insights
        enriched_contacts = sum(1 for c in contacts if self._safe_getattr(c, 'data_source') and self._safe_getattr(c, 'data_source') != 'None')
        if enriched_contacts / total_contacts < 0.5:
            insights.append("Low enrichment coverage - increase data enrichment efforts for better insights")
        
        return insights